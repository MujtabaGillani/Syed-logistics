"""Excel (.xlsx) import helpers for customers, vouchers and expenses.

Each importer reads the first worksheet, treats row 1 as a header row (matched
case-insensitively against known column names) and creates records row by row.
Every importer returns a summary dict:

    {"created": int, "updated": int, "skipped": int, "errors": [{"row", "error"}]}

Imports are transactional per row — a bad row is reported and skipped without
aborting the whole file.
"""
from datetime import datetime, date, time
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

from .models import Customer, GeneralVoucher, OfficeExpense


def _norm(value):
    return str(value).strip().lower() if value is not None else ''


def _read_rows(file_obj):
    """Yield (row_number, {normalized_header: value}) for each data row."""
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header = next(rows)
    except StopIteration:
        return
    headers = [_norm(h) for h in header]
    for i, row in enumerate(rows, start=2):
        if all(c is None or str(c).strip() == '' for c in row):
            continue  # blank line
        record = {}
        for h, value in zip(headers, row):
            if h:
                record[h] = value
        yield i, record


def _get(record, *names):
    for n in names:
        if n in record and record[n] not in (None, ''):
            return record[n]
    return None


def _to_decimal(value, field='amount'):
    if value in (None, ''):
        raise ValueError(f'{field} is required')
    try:
        d = Decimal(str(value).replace(',', '').strip())
    except (InvalidOperation, ValueError):
        raise ValueError(f'{field} "{value}" is not a valid number')
    if d < 0:
        raise ValueError(f'{field} cannot be negative')
    return d


def _to_date(value, field='date', required=True):
    if value in (None, ''):
        if required:
            raise ValueError(f'{field} is required')
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%m/%d/%Y'):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f'{field} "{value}" is not a valid date (use YYYY-MM-DD)')


def _to_time(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.time()
    if isinstance(value, time):
        return value
    text = str(value).strip()
    for fmt in ('%H:%M', '%H:%M:%S', '%I:%M %p'):
        try:
            return datetime.strptime(text, fmt).time()
        except ValueError:
            continue
    return None


def _match_choice(value, choices, default=None):
    """Match a cell against (value, label) choice pairs, by value or label."""
    if value in (None, ''):
        return default
    n = _norm(value)
    for val, label in choices:
        if n == _norm(val) or n == _norm(label):
            return val
    return default


# --------------------------------------------------------------------------
def import_customers(file_obj):
    result = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
    for rownum, rec in _read_rows(file_obj):
        try:
            name = _get(rec, 'name', 'first name')
            sur_name = _get(rec, 'surname', 'sur_name', 'last name') or ''
            cnic = _get(rec, 'cnic')
            contact = _get(rec, 'contact', 'contact number', 'contact_number')
            city = _get(rec, 'city')
            address = _get(rec, 'address') or ''
            if not name or not cnic or not contact or not city:
                raise ValueError('name, cnic, contact and city are required')
            category = _match_choice(
                _get(rec, 'category', 'customer category'),
                Customer.CATEGORY_CHOICES, Customer.CATEGORY_RETAIL)
            defaults = {
                'name': str(name).strip(),
                'sur_name': str(sur_name).strip(),
                'contact_number': str(contact).strip(),
                'city': str(city).strip(),
                'address': str(address).strip(),
                'email': (_get(rec, 'email') or None),
                'customer_category': category,
            }
            obj, created = Customer.objects.update_or_create(
                cnic=str(cnic).strip(), defaults=defaults)
            result['created' if created else 'updated'] += 1
        except Exception as exc:  # noqa: BLE001 - report row-level errors
            result['errors'].append({'row': rownum, 'error': str(exc)})
            result['skipped'] += 1
    return result


def _find_customer(rec):
    cnic = _get(rec, 'cnic', 'customer cnic')
    if cnic:
        c = Customer.objects.filter(cnic=str(cnic).strip()).first()
        if c:
            return c
    name = _get(rec, 'customer', 'customer name')
    if name:
        name = str(name).strip()
        # Try "First Last" then first-name match.
        for c in Customer.objects.all():
            if c.full_name.lower() == name.lower():
                return c
        first = name.split(' ')[0]
        c = Customer.objects.filter(name__iexact=first).first()
        if c:
            return c
    return None


def import_vouchers(file_obj):
    result = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
    for rownum, rec in _read_rows(file_obj):
        try:
            invoice_no = _get(rec, 'invoice #', 'invoice', 'invoice number',
                              'invoice_number')
            # Invoice number is optional — auto-generated when omitted.
            invoice_no = str(invoice_no).strip() if invoice_no else ''
            if invoice_no and GeneralVoucher.objects.filter(
                    invoice_number=invoice_no).exists():
                result['skipped'] += 1
                result['errors'].append(
                    {'row': rownum,
                     'error': f'invoice {invoice_no} already exists — skipped'})
                continue
            customer = _find_customer(rec)
            if not customer:
                raise ValueError(
                    'customer not found (match by CNIC or exact full name)')
            amount = _to_decimal(_get(rec, 'amount'), 'amount')
            invoice_date = _to_date(
                _get(rec, 'date', 'invoice date', 'invoice_date'),
                'invoice date')
            due_date = _to_date(_get(rec, 'due date', 'due_date'),
                                'due date', required=False)
            payment_type = _match_choice(
                _get(rec, 'payment type', 'payment_type', 'type'),
                GeneralVoucher.PAYMENT_CHOICES, GeneralVoucher.PAYMENT_CASH)
            voucher = GeneralVoucher.objects.create(
                invoice_number=invoice_no,
                invoice_date=invoice_date,
                customer=customer,
                payment_type=payment_type,
                amount=amount,
                due_date=due_date,
                notes=str(_get(rec, 'notes') or '').strip(),
            )
            voucher.recompute_paid()
            result['created'] += 1
        except Exception as exc:  # noqa: BLE001
            result['errors'].append({'row': rownum, 'error': str(exc)})
            result['skipped'] += 1
    return result


def import_expenses(file_obj):
    result = {'created': 0, 'updated': 0, 'skipped': 0, 'errors': []}
    for rownum, rec in _read_rows(file_obj):
        try:
            name = _get(rec, 'name')
            if not name:
                raise ValueError('name is required')
            amount = _to_decimal(_get(rec, 'amount'), 'amount')
            exp_date = _to_date(_get(rec, 'date'), 'date')
            expense_type = _match_choice(
                _get(rec, 'type', 'expense type', 'expense_type'),
                OfficeExpense.TYPE_CHOICES, OfficeExpense.TYPE_OTHER)
            OfficeExpense.objects.create(
                name=str(name).strip(),
                amount=amount,
                date=exp_date,
                time=_to_time(_get(rec, 'time')),
                expense_type=expense_type,
                notes=str(_get(rec, 'notes') or '').strip(),
            )
            result['created'] += 1
        except Exception as exc:  # noqa: BLE001
            result['errors'].append({'row': rownum, 'error': str(exc)})
            result['skipped'] += 1
    return result
