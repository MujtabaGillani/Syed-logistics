"""Export helpers: render finance data to Excel (.xlsx) and PDF.

Excel uses openpyxl; PDF uses reportlab. Each function returns a ready-to-send
``HttpResponse`` with the right content-type and download filename.
"""
from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
)

HEADER_FILL = PatternFill('solid', fgColor='06A3DA')
HEADER_FONT = Font(bold=True, color='FFFFFF')


def _stamp():
    return timezone.localtime().strftime('%Y%m%d-%H%M')


def _money(value):
    try:
        return f'{Decimal(value):,.2f}'
    except Exception:
        return value


# --------------------------------------------------------------------------
# Excel
# --------------------------------------------------------------------------
def _excel_response(title, headers, rows, filename, numeric_cols=()):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    for col, head in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=head)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center')

    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if c in numeric_cols and isinstance(value, (int, float, Decimal)):
                cell.number_format = '#,##0.00'

    # Auto-ish column widths.
    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        width = max(
            [len(str(headers[col - 1]))]
            + [len(str(row[col - 1])) for row in rows if col - 1 < len(row)]
            + [10]
        )
        ws.column_dimensions[letter].width = min(width + 2, 45)

    ws.freeze_panes = 'A2'

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    resp = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}-{_stamp()}.xlsx"'
    return resp


# --------------------------------------------------------------------------
# PDF
# --------------------------------------------------------------------------
def _pdf_response(heading, headers, rows, filename, totals_row=None,
                  col_widths=None, landscape_mode=True):
    buffer = BytesIO()
    page = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(
        buffer, pagesize=page,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=14 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    elements = [
        Paragraph(f'<b>Syed Logistic</b>', styles['Title']),
        Paragraph(heading, styles['Heading2']),
        Paragraph(
            f'Generated {timezone.localtime().strftime("%d %b %Y, %H:%M")}',
            styles['Normal']),
        Spacer(1, 8),
    ]

    data = [headers] + rows
    if totals_row:
        data.append(totals_row)

    table = Table(data, repeatRows=1, colWidths=col_widths)
    style = [
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#06A3DA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DCE5')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#F4F6FB')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]
    if totals_row:
        style.append(('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'))
        style.append(('BACKGROUND', (0, -1), (-1, -1),
                      colors.HexColor('#E9EEF6')))
    table.setStyle(TableStyle(style))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}-{_stamp()}.pdf"'
    return resp


# --------------------------------------------------------------------------
# Vouchers
# --------------------------------------------------------------------------
VOUCHER_HEADERS = ['Invoice #', 'Date', 'Customer', 'Payment Type', 'Amount',
                   'Paid', 'Outstanding', 'Due Date', 'Status']


def _voucher_row(v):
    return [
        v.invoice_number,
        v.invoice_date.isoformat() if v.invoice_date else '',
        v.customer.full_name,
        v.get_payment_type_display(),
        v.signed_amount,
        v.total_paid,
        v.outstanding,
        v.due_date.isoformat() if v.due_date else '',
        'Settled' if v.is_paid else 'Due',
    ]


def vouchers_excel(rows):
    data = [_voucher_row(v) for v in rows]
    return _excel_response('Vouchers', VOUCHER_HEADERS, data,
                           'vouchers', numeric_cols=(5, 6, 7))


def vouchers_pdf(rows):
    data = []
    total_amt = total_paid = total_out = Decimal('0.00')
    for v in rows:
        data.append([
            v.invoice_number,
            v.invoice_date.isoformat() if v.invoice_date else '',
            v.customer.full_name,
            v.get_payment_type_display(),
            _money(v.signed_amount),
            _money(v.total_paid),
            _money(v.outstanding),
            'Settled' if v.is_paid else 'Due',
        ])
        total_amt += v.signed_amount
        total_paid += v.total_paid
        total_out += v.outstanding
    headers = ['Invoice #', 'Date', 'Customer', 'Type', 'Amount', 'Paid',
               'Outstanding', 'Status']
    totals = ['', '', '', 'TOTAL', _money(total_amt), _money(total_paid),
              _money(total_out), '']
    return _pdf_response('General Vouchers', headers, data, 'vouchers',
                         totals_row=totals)


# --------------------------------------------------------------------------
# Expenses
# --------------------------------------------------------------------------
EXPENSE_HEADERS = ['Name', 'Type', 'Amount', 'Date', 'Time', 'Notes']


def _expense_row(x):
    return [
        x.name,
        x.get_expense_type_display(),
        x.amount,
        x.date.isoformat() if x.date else '',
        x.time.strftime('%H:%M') if x.time else '',
        x.notes,
    ]


def expenses_excel(rows):
    data = [_expense_row(x) for x in rows]
    return _excel_response('Expenses', EXPENSE_HEADERS, data,
                           'expenses', numeric_cols=(3,))


def expenses_pdf(rows):
    data = []
    total = Decimal('0.00')
    for x in rows:
        data.append([
            x.name, x.get_expense_type_display(), _money(x.amount),
            x.date.isoformat() if x.date else '',
            x.time.strftime('%H:%M') if x.time else '',
        ])
        total += x.amount
    headers = ['Name', 'Type', 'Amount', 'Date', 'Time']
    totals = ['', 'TOTAL', _money(total), '', '']
    return _pdf_response('Office Expenses', headers, data, 'expenses',
                         totals_row=totals, landscape_mode=False)


# --------------------------------------------------------------------------
# Customers
# --------------------------------------------------------------------------
CUSTOMER_HEADERS = ['Name', 'Surname', 'CNIC', 'Contact', 'City', 'Category',
                    'Email', 'Address']


def _customer_row(c):
    return [
        c.name, c.sur_name, c.cnic, c.contact_number, c.city,
        c.get_customer_category_display(), c.email or '', c.address,
    ]


def customers_excel(rows):
    data = [_customer_row(c) for c in rows]
    return _excel_response('Customers', CUSTOMER_HEADERS, data, 'customers')


def customers_pdf(rows):
    data = [[c.name, c.sur_name, c.cnic, c.contact_number, c.city,
             c.get_customer_category_display(), c.email or '']
            for c in rows]
    headers = ['Name', 'Surname', 'CNIC', 'Contact', 'City', 'Category',
               'Email']
    return _pdf_response('Customers', headers, data, 'customers')


# --------------------------------------------------------------------------
# Customer statement / ledger
# --------------------------------------------------------------------------
def customer_statement_excel(customer, ledger):
    headers = ['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    rows = []
    for ln in ledger['lines']:
        rows.append([
            ln['date'], ln['reference'], ln['description'],
            Decimal(ln['debit']) if ln['debit'] is not None else '',
            Decimal(ln['credit']) if ln['credit'] is not None else '',
            Decimal(ln['balance']),
        ])
    t = ledger['totals']
    rows.append(['', '', 'TOTAL', Decimal(t['debit']), Decimal(t['credit']),
                 Decimal(t['balance'])])
    return _excel_response(
        f'Statement {customer.full_name}'[:31], headers, rows,
        f'statement-{customer.id}', numeric_cols=(4, 5, 6))


def customer_statement_pdf(customer, ledger):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=14 * mm, bottomMargin=12 * mm,
    )
    styles = getSampleStyleSheet()
    t = ledger['totals']
    elements = [
        Paragraph('<b>Syed Logistic</b>', styles['Title']),
        Paragraph('Customer Statement', styles['Heading2']),
        Paragraph(f"<b>{customer.full_name}</b> &nbsp;|&nbsp; CNIC: "
                  f"{customer.cnic} &nbsp;|&nbsp; {customer.city} "
                  f"&nbsp;|&nbsp; {customer.contact_number}", styles['Normal']),
        Paragraph(
            f'Generated {timezone.localtime().strftime("%d %b %Y, %H:%M")}',
            styles['Normal']),
        Spacer(1, 10),
    ]

    headers = ['Date', 'Reference', 'Description', 'Debit', 'Credit', 'Balance']
    data = [headers]
    for ln in ledger['lines']:
        data.append([
            ln['date'], ln['reference'], ln['description'],
            _money(ln['debit']) if ln['debit'] is not None else '',
            _money(ln['credit']) if ln['credit'] is not None else '',
            _money(ln['balance']),
        ])
    data.append(['', '', 'TOTAL', _money(t['debit']), _money(t['credit']),
                 _money(t['balance'])])

    table = Table(data, repeatRows=1,
                  colWidths=[22 * mm, 28 * mm, 62 * mm, 24 * mm, 24 * mm, 26 * mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#06A3DA')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#D5DCE5')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2),
         [colors.white, colors.HexColor('#F4F6FB')]),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E9EEF6')),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 8))
    elements.append(Paragraph(
        f"<b>Closing balance: Rs {_money(t['balance'])}</b> "
        f"(positive = customer owes us)", styles['Normal']))

    doc.build(elements)
    buffer.seek(0)
    resp = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    resp['Content-Disposition'] = \
        f'attachment; filename="statement-{customer.id}-{_stamp()}.pdf"'
    return resp
